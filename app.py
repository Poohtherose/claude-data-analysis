"""
SPSS-style One-Way ANOVA Analysis Web Application
Supports Levene's test, LSD, and Duncan's Multiple Range Test
"""

from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
import numpy as np
from scipy import stats
from statsmodels.stats.multicomp import pairwise_tukeyhsd
import io
import os
import base64
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from matplotlib import rcParams
import tempfile
import atexit
import shutil
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def levene_test(data_dict):
    """
    方差齐性检验 (Levene's Test)
    类似于SPSS中的方差齐性检验
    """
    groups = [values for values in data_dict.values() if len(values) > 0]
    if len(groups) < 2:
        return None

    statistic, p_value = stats.levene(*groups, center='mean')

    return {
        'statistic': round(statistic, 4),
        'p_value': round(p_value, 4),
        'significant': p_value < 0.05,
        'interpretation': '方差不齐' if p_value < 0.05 else '方差齐性'
    }


def one_way_anova(data_dict):
    """
    单因素ANOVA检验
    返回类似SPSS的ANOVA表格结果
    """
    groups = [values for values in data_dict.values() if len(values) > 0]
    group_names = list(data_dict.keys())

    if len(groups) < 2:
        return None

    # 执行ANOVA
    f_stat, p_value = stats.f_oneway(*groups)

    # 计算各组统计量
    group_stats = []
    all_values = []

    for name, values in data_dict.items():
        if len(values) > 0:
            values_array = np.array(values)
            group_stats.append({
                'group': name,
                'n': len(values),
                'mean': round(np.mean(values_array), 4),
                'std': round(np.std(values_array, ddof=1), 4),
                'se': round(np.std(values_array, ddof=1) / np.sqrt(len(values)), 4),
                'min': round(np.min(values_array), 4),
                'max': round(np.max(values_array), 4)
            })
            all_values.extend(values)

    # 计算总体的统计量
    total_n = len(all_values)
    grand_mean = np.mean(all_values)

    # 计算平方和
    ss_between = sum(len(data_dict[g]) * (np.mean(data_dict[g]) - grand_mean) ** 2
                     for g in data_dict if len(data_dict[g]) > 0)
    ss_within = sum(sum((x - np.mean(data_dict[g])) ** 2 for x in data_dict[g])
                    for g in data_dict if len(data_dict[g]) > 0)
    ss_total = sum((x - grand_mean) ** 2 for g in data_dict for x in data_dict[g] if len(data_dict[g]) > 0)

    df_between = len(groups) - 1
    df_within = total_n - len(groups)
    df_total = total_n - 1

    ms_between = ss_between / df_between if df_between > 0 else 0
    ms_within = ss_within / df_within if df_within > 0 else 0

    # 重新计算F值以确保一致性
    f_stat = ms_between / ms_within if ms_within > 0 else 0

    return {
        'f_statistic': round(f_stat, 4),
        'p_value': round(p_value, 4),
        'significant': p_value < 0.05,
        'group_stats': group_stats,
        'anova_table': {
            'between_groups': {
                'ss': round(ss_between, 4),
                'df': df_between,
                'ms': round(ms_between, 4),
                'f': round(f_stat, 4),
                'p': round(p_value, 4)
            },
            'within_groups': {
                'ss': round(ss_within, 4),
                'df': df_within,
                'ms': round(ms_within, 4)
            },
            'total': {
                'ss': round(ss_total, 4),
                'df': df_total
            }
        }
    }


def lsd_test(data_dict, anova_result, alpha=0.05):
    """
    LSD (Least Significant Difference) 检验
    类似SPSS中的LSD多重比较
    """
    groups = list(data_dict.keys())
    n_groups = len(groups)

    if n_groups < 2:
        return None

    # 获取MS_within和df_within
    ms_within = anova_result['anova_table']['within_groups']['ms']
    df_within = anova_result['anova_table']['within_groups']['df']

    # 计算每组的均值和样本量
    group_means = {}
    group_ns = {}

    for group in groups:
        values = data_dict[group]
        if len(values) > 0:
            group_means[group] = np.mean(values)
            group_ns[group] = len(values)

    # 计算LSD值
    # LSD = t_(alpha/2, df_within) * sqrt(MS_within * (1/ni + 1/nj))
    t_critical = stats.t.ppf(1 - alpha / 2, df_within)

    results = []
    group_list = list(group_means.keys())

    for i in range(len(group_list)):
        for j in range(i + 1, len(group_list)):
            group1 = group_list[i]
            group2 = group_list[j]

            mean1 = group_means[group1]
            mean2 = group_means[group2]
            n1 = group_ns[group1]
            n2 = group_ns[group2]

            mean_diff = mean1 - mean2
            std_error = np.sqrt(ms_within * (1 / n1 + 1 / n2))
            t_stat = mean_diff / std_error if std_error > 0 else 0
            p_value = 2 * (1 - stats.t.cdf(abs(t_stat), df_within))

            # 计算此比较的LSD值
            lsd_value = t_critical * std_error

            results.append({
                'group1': group1,
                'group2': group2,
                'mean1': round(mean1, 4),
                'mean2': round(mean2, 4),
                'mean_diff': round(mean_diff, 4),
                'std_error': round(std_error, 4),
                't_stat': round(t_stat, 4),
                'p_value': round(p_value, 4),
                'lsd_value': round(lsd_value, 4),
                'significant': abs(mean_diff) > lsd_value,
                'alpha': alpha
            })

    return {
        'lsd_value': round(lsd_value, 4),
        'alpha': alpha,
        'comparisons': results
    }


def duncan_test(data_dict, anova_result, alpha=0.05):
    """
    Duncan's Multiple Range Test (邓肯检验)
    类似SPSS中的Duncan检验
    """
    groups = list(data_dict.keys())
    n_groups = len(groups)

    if n_groups < 2:
        return None

    # 获取ANOVA结果
    ms_within = anova_result['anova_table']['within_groups']['ms']
    df_within = anova_result['anova_table']['within_groups']['df']

    # 计算每组的均值和样本量
    group_stats = []
    for group in groups:
        values = data_dict[group]
        if len(values) > 0:
            group_stats.append({
                'group': group,
                'mean': np.mean(values),
                'n': len(values)
            })

    # 按均值降序排序
    group_stats.sort(key=lambda x: x['mean'], reverse=True)

    # 计算标准误 (基于平均样本量或各组样本量)
    # 使用调和平均数处理不等样本量
    ns = [gs['n'] for gs in group_stats]
    n_harmonic = len(ns) / sum(1 / n for n in ns) if all(n > 0 for n in ns) else np.mean(ns)
    standard_error = np.sqrt(ms_within / n_harmonic)

    # 使用t分布近似计算Duncan临界值（比studentized_range快得多）
    # Duncan检验使用变化的显著性水平: alpha^(k-1) 其中k是跨度
    def get_duncan_critical_range(p, df, alpha, se):
        """使用t分布计算Duncan临界值"""
        # Duncan检验: 对p个均值，使用 alpha_p = 1 - (1-alpha)^(p-1)
        alpha_p = 1 - (1 - alpha) ** (p - 1)
        # 使用t分布的临界值近似
        t_val = stats.t.ppf(1 - alpha_p / 2, df)
        # 转换为范围统计量
        return t_val * se * np.sqrt(2)

    # 计算各组间的显著性
    comparisons = []
    k = len(group_stats)

    for i in range(k):
        for j in range(i + 1, k):
            group1 = group_stats[i]
            group2 = group_stats[j]

            mean_diff = abs(group1['mean'] - group2['mean'])
            p = j - i + 1  # 跨度 (number of means spanned)

            # 计算Duncan临界值
            critical_range = get_duncan_critical_range(p, df_within, alpha, standard_error)

            # 使用t检验计算p值（更快且结果接近）
            t_stat = mean_diff / (standard_error * np.sqrt(2)) if standard_error > 0 else 0
            p_value = 2 * (1 - stats.t.cdf(abs(t_stat), df_within))

            comparisons.append({
                'group1': group1['group'],
                'group2': group2['group'],
                'mean1': round(group1['mean'], 4),
                'mean2': round(group2['mean'], 4),
                'mean_diff': round(mean_diff, 4),
                'span': p,
                'critical_range': round(critical_range, 4),
                'significant': mean_diff > critical_range,
                'p_value': round(p_value, 4)
            })

    # 为每个组分配字母标记（类似SPSS输出）
    # 这是一个简化的实现
    group_markers = {gs['group']: '' for gs in group_stats}
    current_letter = 'a'

    for i, gs in enumerate(group_stats):
        if group_markers[gs['group']] == '':
            # 为该组及其不显著不同的组分配相同的字母
            group_markers[gs['group']] = current_letter

            for j in range(i + 1, len(group_stats)):
                other = group_stats[j]
                # 检查是否显著不同
                comp = next((c for c in comparisons
                           if (c['group1'] == gs['group'] and c['group2'] == other['group']) or
                              (c['group1'] == other['group'] and c['group2'] == gs['group'])), None)

                if comp and not comp['significant'] and group_markers[other['group']] == '':
                    group_markers[other['group']] = current_letter

            current_letter = chr(ord(current_letter) + 1)

    return {
        'standard_error': round(standard_error, 4),
        'ms_within': round(ms_within, 4),
        'df_within': df_within,
        'alpha': alpha,
        'group_means': [{**gs, 'mean': round(gs['mean'], 4), 'subset': group_markers[gs['group']]} for gs in group_stats],
        'comparisons': comparisons
    }


def merge_parallel_samples(data_dict, suffix_length=1):
    """
    将平行样合并为组
    例如: {'0K1': [1,2,3], '0K2': [4,5,6]} -> {'0K': [1,2,3,4,5,6]}
    """
    merged = {}
    for sample_name, values in data_dict.items():
        # 去掉后缀
        group_name = sample_name[:-suffix_length] if len(sample_name) > suffix_length else sample_name
        if group_name not in merged:
            merged[group_name] = []
        merged[group_name].extend(values)
    return merged


def detect_column_types(df):
    """
    智能识别列类型
    返回: (sample_column, indicator_columns)
    - sample_column: 样品名称列（字符串类型、重复值较多的列）
    - indicator_columns: 指标列（数值类型列，支持数字、英文、中文列名）
    """
    columns = df.columns.tolist()

    # 分析每列的特征
    column_analysis = []

    for col in columns:
        series = df[col]
        unique_count = series.nunique()
        total_count = len(series)
        unique_ratio = unique_count / total_count if total_count > 0 else 0

        # 检查数据类型
        is_numeric = pd.api.types.is_numeric_dtype(series)
        is_string = pd.api.types.is_string_dtype(series) or pd.api.types.is_object_dtype(series)

        # 计算非空值数量
        non_null_count = series.notna().sum()

        # 样品列的特征：
        # 1. 通常是字符串类型（或混合类型）
        # 2. 重复值较多（unique_ratio 较低）
        # 3. 不全是数值
        # 4. 常见列名：样品、样本、名称、组、group、sample 等
        sample_keywords = ['样品', '样本', '名称', '组', 'group', 'sample', 'name', 'id', '编号', '处理', 'treatment', 'sample', 'id']
        has_sample_keyword = any(keyword in str(col).lower() for keyword in sample_keywords)

        # 指标列的特征：
        # 1. 数值类型
        # 2. 通常是连续的数值
        # 3. 列名可以是纯数字、英文、中文或混合
        indicator_keywords = ['指标', '结果', '含量', '活性', '值', 'value', 'result', 'concentration', 'activity', 'level', 'data', 'content', 'ph', 'ratio', 'rate', 'time', 'temp', 'weight', 'mass']
        has_indicator_keyword = any(keyword in str(col).lower() for keyword in indicator_keywords)

        # 检查列名是否是纯数字（如300, 310, 320）- 这也可能是指标列
        col_str = str(col).strip()
        is_pure_numeric_name = col_str.replace('.', '').replace('-', '').isdigit()

        column_analysis.append({
            'column': col,
            'is_numeric': is_numeric,
            'is_string': is_string,
            'unique_count': unique_count,
            'unique_ratio': unique_ratio,
            'non_null_count': non_null_count,
            'has_sample_keyword': has_sample_keyword,
            'has_indicator_keyword': has_indicator_keyword,
            'is_pure_numeric_name': is_pure_numeric_name,
            'col_str': col_str
        })

    # 识别样品列
    sample_candidates = []
    for analysis in column_analysis:
        score = 0

        # 优先根据关键词匹配
        if analysis['has_sample_keyword']:
            score += 100

        # 字符串类型加分
        if analysis['is_string']:
            score += 20

        # 重复值较多加分（但不是唯一值）
        if 0.1 < analysis['unique_ratio'] < 0.8:
            score += 15
        elif analysis['unique_ratio'] <= 0.1:
            score += 10

        # 非数值类型加分
        if not analysis['is_numeric']:
            score += 10

        # 第一列通常可能是样品列
        if columns.index(analysis['column']) == 0:
            score += 5

        # 唯一值数量适中（2-50个组）
        if 2 <= analysis['unique_count'] <= 50:
            score += 10

        # 纯数字列名不太可能是样品名称（更可能是指标）
        if analysis['is_pure_numeric_name']:
            score -= 50

        sample_candidates.append((analysis['column'], score, analysis))

    # 按分数排序，选择最可能是样品列的
    # 使用 str(x[0]) 作为次要排序键，避免 int 和 str 比较错误
    sample_candidates.sort(key=lambda x: (x[1], str(x[0])), reverse=True)
    sample_column = sample_candidates[0][0] if sample_candidates else columns[0]

    # 识别指标列：所有数值列，排除样品列
    # 支持纯数字列名、英文列名、中文列名或混合
    indicator_columns = []
    for analysis in column_analysis:
        if analysis['column'] == sample_column:
            continue

        # 如果列数据是数值类型，就认为是指标列
        if analysis['is_numeric'] and analysis['non_null_count'] > 0:
            indicator_columns.append(analysis['column'])
            continue

        # 如果列名是纯数字（如300, 310），尝试转换为数值
        if analysis['is_pure_numeric_name'] and not analysis['is_numeric']:
            try:
                converted = pd.to_numeric(df[analysis['column']], errors='coerce')
                if converted.notna().sum() > 0:
                    indicator_columns.append(analysis['column'])
            except:
                pass

    # 如果没有找到数值列，检查是否有其他列可以转换
    if not indicator_columns:
        for analysis in column_analysis:
            if analysis['column'] == sample_column:
                continue
            # 尝试转换为数值
            try:
                converted = pd.to_numeric(df[analysis['column']], errors='coerce')
                if converted.notna().sum() > 0:
                    indicator_columns.append(analysis['column'])
            except:
                pass

    return sample_column, indicator_columns


def find_column_name(df, column_name_str):
    """
    根据字符串列名找到DataFrame中实际的列名（处理int/str类型差异）
    例如：df列名为300(int)，传入"300"(str)，返回300(int)
    """
    if column_name_str is None:
        return None
    # 先直接尝试
    if column_name_str in df.columns:
        return column_name_str
    # 尝试匹配字符串形式
    for col in df.columns:
        if str(col) == str(column_name_str):
            return col
    return column_name_str  # 如果没找到，返回原始值让后续报错


def process_data(df, sample_column=None, value_columns=None, merge_parallel=False, merge_suffix_length=1):
    """
    处理上传的Excel数据
    自动识别或使用指定的列
    value_columns: 可以是单个列名或列名列表
    """
    # 如果未指定列，尝试自动识别
    if sample_column is None or value_columns is None:
        # 智能识别列类型
        detected_sample, detected_indicators = detect_column_types(df)

        if sample_column is None:
            sample_column = detected_sample

        # 如果用户指定了指标列，使用用户指定的
        if value_columns is None:
            # 如果没有指定，使用所有检测到的指标列
            if len(detected_indicators) == 0:
                raise ValueError(f"未找到数值型指标列。请确保数据文件中包含数值数据（样品列：{sample_column}）")
            selected_indicators = detected_indicators
        else:
            # 处理单个列名或列表
            if isinstance(value_columns, str):
                # 可能是逗号分隔的多个列名
                if ',' in value_columns:
                    value_columns = [c.strip() for c in value_columns.split(',')]
                else:
                    value_columns = [value_columns]

            # 将用户传入的列名（字符串）转换为DataFrame中实际的列名
            selected_indicators = []
            for col in value_columns:
                actual_col = find_column_name(df, col)
                if actual_col in df.columns:
                    selected_indicators.append(actual_col)

            # 如果没有找到有效的指标列，使用检测到的所有指标列
            if not selected_indicators:
                selected_indicators = detected_indicators

            # 验证选中的列是否在检测到的指标列中
            valid_indicators = []
            for col in selected_indicators:
                if col in detected_indicators:
                    valid_indicators.append(col)
                elif pd.api.types.is_numeric_dtype(df[col]):
                    valid_indicators.append(col)
            selected_indicators = valid_indicators if valid_indicators else detected_indicators
    else:
        # 处理单个列名或列表
        if isinstance(value_columns, str):
            if ',' in value_columns:
                selected_indicators = [find_column_name(df, c.strip()) for c in value_columns.split(',')]
            else:
                selected_indicators = [find_column_name(df, value_columns)]
        else:
            selected_indicators = [find_column_name(df, c) for c in value_columns]

        # 过滤掉无效的列
        selected_indicators = [c for c in selected_indicators if c in df.columns]

    # 将样品列名转换为正确的类型
    sample_column = find_column_name(df, sample_column)

    # 构建数据字典
    data_dict = {}

    # 如果只选择了一个指标列
    if len(selected_indicators) == 1:
        value_column = selected_indicators[0]
        for sample in df[sample_column].unique():
            sample_data = df[df[sample_column] == sample][value_column].dropna()
            data_dict[str(sample)] = sample_data.values.tolist()

        # 如果启用合并平行样，合并数据
        if merge_parallel:
            data_dict = merge_parallel_samples(data_dict, merge_suffix_length)

        return {value_column: data_dict}, [value_column]
    else:
        # 多指标情况，分别处理每个指标
        results = {}
        for col in selected_indicators:
            data_dict_single = {}
            for sample in df[sample_column].unique():
                sample_data = df[df[sample_column] == sample][col].dropna()
                data_dict_single[str(sample)] = sample_data.values.tolist()
            # 如果启用合并平行样，合并数据
            if merge_parallel:
                data_dict_single = merge_parallel_samples(data_dict_single, merge_suffix_length)
            results[col] = data_dict_single
        return results, selected_indicators


def create_excel_report(all_results, original_df):
    """
    创建类似SPSS输出的Excel报告
    """
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # ---- 汇总表：所有指标的均值、标准差、Duncan分组 ----
    all_samples = []
    for indicator, results in all_results.items():
        for gs in results['anova']['group_stats']:
            if gs['group'] not in all_samples:
                all_samples.append(gs['group'])

    summary_rows = []
    for sample in all_samples:
        row = {'样品名称': sample}
        for indicator, results in all_results.items():
            ind_key = str(indicator)
            gs_match = next((gs for gs in results['anova']['group_stats'] if gs['group'] == sample), None)
            duncan_match = next((gm for gm in results['duncan']['group_means'] if gm['group'] == sample), None)
            row[f'{ind_key}_平均值'] = round(gs_match['mean'], 4) if gs_match else None
            row[f'{ind_key}_标准差'] = round(gs_match['std'], 4) if gs_match else None
            row[f'{ind_key}_Duncan'] = duncan_match['subset'] if duncan_match else ''
        summary_rows.append(row)

    summary_df = pd.DataFrame(summary_rows)
    summary_df.to_excel(writer, sheet_name='汇总', index=False, startrow=0)

    # ---- 各指标详细表 ----
    for indicator, results in all_results.items():
        sheet_name = str(indicator)[:31]  # Excel工作表名称限制31字符

        # 获取数据
        anova_data = results['anova']['anova_table']
        levene = results['levene']
        lsd = results['lsd']
        duncan = results['duncan']
        desc_stats_data = results['anova']['group_stats']

        # 创建描述性统计表 - 使用中文表头和指定列顺序
        # 按均值降序排序，与Duncan检验结果对应
        desc_stats = pd.DataFrame({
            '样品名称': [gs['group'] for gs in desc_stats_data],
            '平均值': [gs['mean'] for gs in desc_stats_data],
            '标准差': [gs['std'] for gs in desc_stats_data],
            '标准误差': [gs['se'] for gs in desc_stats_data]
        })

        # 添加Duncan相关性列 - 从duncan结果中获取
        duncan_subsets = {}
        for gm in duncan['group_means']:
            duncan_subsets[gm['group']] = gm['subset']
        desc_stats['Duncan相关性'] = desc_stats['样品名称'].map(duncan_subsets)

        # 按平均值降序排序（Duncan检验的标准做法）
        desc_stats = desc_stats.sort_values('平均值', ascending=False).reset_index(drop=True)

        # 创建ANOVA表 - 使用中文
        anova_df = pd.DataFrame({
            '变异来源': ['组间', '组内', '总计'],
            '平方和': [anova_data['between_groups']['ss'],
                      anova_data['within_groups']['ss'],
                      anova_data['total']['ss']],
            '自由度': [anova_data['between_groups']['df'],
                     anova_data['within_groups']['df'],
                     anova_data['total']['df']],
            '均方': [anova_data['between_groups']['ms'],
                   anova_data['within_groups']['ms'],
                   ''],
            'F值': [anova_data['between_groups']['f'], '', ''],
            '显著性': [anova_data['between_groups']['p'], '', '']
        })

        # 创建方差齐性检验表 - 使用中文
        levene_df = pd.DataFrame({
            '检验方法': ["Levene方差齐性检验"],
            '统计量': [levene['statistic']],
            '自由度1': [len(results['anova']['group_stats']) - 1],
            '自由度2': [anova_data['within_groups']['df']],
            '显著性': [levene['p_value']],
            '结果解释': [levene['interpretation']]
        })

        # 创建LSD检验表 - 使用中文
        lsd_df = pd.DataFrame(lsd['comparisons'])
        lsd_df = lsd_df[['group1', 'group2', 'mean_diff', 'std_error', 't_stat', 'p_value', 'significant']]
        lsd_df.columns = ['组I', '组J', '均值差(I-J)', '标准误', 't值', '显著性', '是否显著']

        # 创建Duncan检验表 - 使用中文
        duncan_comparisons_df = pd.DataFrame(duncan['comparisons'])
        duncan_comparisons_df = duncan_comparisons_df[['group1', 'group2', 'mean_diff', 'critical_range', 'significant']]
        duncan_comparisons_df.columns = ['组I', '组J', '均值差', '临界范围', '是否显著']

        duncan_means_df = pd.DataFrame(duncan['group_means'])
        duncan_means_df = duncan_means_df[['group', 'n', 'mean', 'subset']]
        duncan_means_df.columns = ['样品名称', '样本量', '平均值', 'Duncan分组']

        # 写入Excel
        start_row = 0

        # 描述性统计
        desc_stats.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
        start_row += len(desc_stats) + 3

        # ANOVA表
        anova_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
        start_row += len(anova_df) + 3

        # 方差齐性检验
        levene_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
        start_row += len(levene_df) + 3

        # LSD检验
        lsd_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
        start_row += len(lsd_df) + 3

        # Duncan检验 - 均值排序
        duncan_means_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
        start_row += len(duncan_means_df) + 3

        # Duncan检验 - 两两比较
        duncan_comparisons_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)

    writer.close()
    output.seek(0)
    return output


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有文件上传'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400

        if not allowed_file(file.filename):
            return jsonify({'error': '只支持 .xlsx 和 .xls 文件'}), 400

        # 读取Excel文件
        df = pd.read_excel(file)
        # 将所有列名统一转为字符串，避免 int/str 混合导致的比较错误
        df.columns = [str(c) for c in df.columns]

        # 获取列选择
        sample_column = request.form.get('sample_column')
        value_columns = request.form.get('value_columns')  # 可以是逗号分隔的多个列名

        # 获取合并平行样参数
        merge_parallel = request.form.get('merge_parallel', 'false').lower() == 'true'
        merge_suffix_length = int(request.form.get('merge_suffix_length', '1') or '1')

        # 转换列名以处理int/str类型差异
        actual_sample_col = find_column_name(df, sample_column) if sample_column else None

        if sample_column and actual_sample_col not in df.columns:
            return jsonify({'error': f'样品列 "{sample_column}" 不存在'}), 400

        # 处理多指标列（逗号分隔）
        if value_columns:
            selected_cols = [c.strip() for c in value_columns.split(',') if c.strip()]
            for col in selected_cols:
                actual_col = find_column_name(df, col)
                if actual_col not in df.columns:
                    return jsonify({'error': f'指标列 "{col}" 不存在'}), 400

        # 使用转换后的列名
        sample_column = actual_sample_col

        # 处理数据
        data_dicts, indicators = process_data(df, sample_column, value_columns, merge_parallel, merge_suffix_length)

        # 对所有指标执行统计分析
        all_results = {}

        for indicator, data_dict in data_dicts.items():
            # 确保有足够的数据组
            valid_groups = {k: v for k, v in data_dict.items() if len(v) > 0}

            if len(valid_groups) < 2:
                continue

            # 执行各项检验
            anova_result = one_way_anova(valid_groups)
            levene_result = levene_test(valid_groups)
            lsd_result = lsd_test(valid_groups, anova_result)
            duncan_result = duncan_test(valid_groups, anova_result)

            all_results[indicator] = {
                'anova': anova_result,
                'levene': levene_result,
                'lsd': lsd_result,
                'duncan': duncan_result
            }

        if not all_results:
            return jsonify({'error': '数据不足以进行ANOVA分析，请确保至少有两个组且每组至少有一个数据点'}), 400

        # 保存结果到session或临时存储（简化处理，直接返回结果）
        # 在实际应用中，可以使用session或Redis等存储

        # 构建汇总表：每个样品 × 每个指标的均值、标准差、Duncan字母
        # 收集所有样品名称（保持顺序）
        all_samples = []
        for indicator, res in all_results.items():
            for gs in res['anova']['group_stats']:
                if gs['group'] not in all_samples:
                    all_samples.append(gs['group'])

        summary_table = []
        for sample in all_samples:
            row = {'sample': sample}
            for indicator, res in all_results.items():
                ind_key = str(indicator)
                # 均值和标准差
                gs_match = next((gs for gs in res['anova']['group_stats'] if gs['group'] == sample), None)
                if gs_match:
                    row[ind_key + '_mean'] = round(gs_match['mean'], 4)
                    row[ind_key + '_std'] = round(gs_match['std'], 4)
                else:
                    row[ind_key + '_mean'] = None
                    row[ind_key + '_std'] = None
                # Duncan字母
                duncan_match = next((gm for gm in res['duncan']['group_means'] if gm['group'] == sample), None)
                row[ind_key + '_duncan'] = duncan_match['subset'] if duncan_match else ''
            summary_table.append(row)

        return jsonify({
            'success': True,
            'indicators': list(all_results.keys()),
            'summary_table': summary_table,
            'preview': {k: {
                'anova_significant': bool(v['anova']['significant']),
                'f_statistic': float(v['anova']['f_statistic']),
                'p_value': float(v['anova']['p_value']),
                'levene_significant': bool(v['levene']['significant']) if v['levene'] else None
            } for k, v in all_results.items()}
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/download_report', methods=['POST'])
def download_report():
    try:
        # 从请求中获取分析结果
        # 注意：在实际应用中，应该从session或数据库获取
        # 这里为了简化，重新执行分析

        if 'file' not in request.files:
            return jsonify({'error': '没有文件上传'}), 400

        file = request.files['file']
        df = pd.read_excel(file)
        # 将所有列名统一转为字符串，避免 int/str 混合导致的比较错误
        df.columns = [str(c) for c in df.columns]

        sample_column = request.form.get('sample_column')
        value_columns = request.form.get('value_columns')  # 可以是逗号分隔的多个列名

        # 获取合并平行样参数
        merge_parallel = request.form.get('merge_parallel', 'false').lower() == 'true'
        merge_suffix_length = int(request.form.get('merge_suffix_length', '1') or '1')

        data_dicts, indicators = process_data(df, sample_column, value_columns, merge_parallel, merge_suffix_length)

        all_results = {}
        for indicator, data_dict in data_dicts.items():
            valid_groups = {k: v for k, v in data_dict.items() if len(v) > 0}

            if len(valid_groups) < 2:
                continue

            anova_result = one_way_anova(valid_groups)
            levene_result = levene_test(valid_groups)
            lsd_result = lsd_test(valid_groups, anova_result)
            duncan_result = duncan_test(valid_groups, anova_result)

            all_results[indicator] = {
                'anova': anova_result,
                'levene': levene_result,
                'lsd': lsd_result,
                'duncan': duncan_result
            }

        excel_file = create_excel_report(all_results, df)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='ANOVA_分析报告.xlsx'
        )

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/get_columns', methods=['POST'])
def get_columns():
    """获取Excel文件的列名，并智能识别列类型"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有文件上传'}), 400

        file = request.files['file']
        if not allowed_file(file.filename):
            return jsonify({'error': '不支持的文件格式'}), 400

        df = pd.read_excel(file)
        # 将所有列名统一转为字符串，避免 int/str 混合导致的比较错误
        df.columns = [str(c) for c in df.columns]

        # 智能识别列类型
        detected_sample, detected_indicators = detect_column_types(df)

        # 将列名转换为字符串，避免 int/str 混合导致的排序/比较错误
        columns_str = [str(c) for c in df.columns]
        detected_sample_str = str(detected_sample)
        detected_indicators_str = [str(c) for c in detected_indicators]

        # 分析每列的详细信息
        column_info = []
        for i, col in enumerate(df.columns):
            series = df[col]
            is_numeric = pd.api.types.is_numeric_dtype(series)
            unique_count = series.nunique()

            column_info.append({
                'name': columns_str[i],
                'type': 'numeric' if is_numeric else 'string',
                'unique_count': int(unique_count),
                'is_indicator': columns_str[i] in detected_indicators_str,
                'is_sample_candidate': columns_str[i] == detected_sample_str
            })

        return jsonify({
            'columns': columns_str,
            'preview': df.head(10).to_dict(orient='records'),
            'all_data': df.to_dict(orient='records'),
            'shape': df.shape,
            'detected': {
                'sample_column': detected_sample_str,
                'indicator_columns': detected_indicators_str
            },
            'column_info': column_info
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


def strip_stat_suffix(col_name):
    """去掉 _mean / _std / _duncan / _sem 等统计后缀，仅保留指标名"""
    s = str(col_name)
    for suffix in ('_mean', '_std', '_duncan', '_sem', '_sd'):
        if s.endswith(suffix):
            return s[:-len(suffix)]
    return s


def setup_fonts():
    """配置字体：中文宋体，英文/数字/标点 Times New Roman"""
    import os, urllib.request

    base_dir = os.path.dirname(os.path.abspath(__file__))
    fonts_dir = os.path.join(base_dir, 'static', 'fonts')
    os.makedirs(fonts_dir, exist_ok=True)

    # --- 宋体 (SimSun) ---
    simsun_path = os.path.join(fonts_dir, 'simsun.ttc')
    if not os.path.exists(simsun_path):
        # 尝试从系统路径复制
        sys_candidates = [
            '/usr/share/fonts/truetype/arphic/uming.ttc',
            '/usr/share/fonts/opentype/noto/NotoSerifCJK-Regular.ttc',
        ]
        copied = False
        for p in sys_candidates:
            if os.path.exists(p):
                import shutil
                shutil.copy2(p, simsun_path)
                copied = True
                break
        if not copied:
            # 下载开源宋体替代（文泉驿正黑 / Noto Serif CJK）
            try:
                url = 'https://github.com/googlefonts/noto-cjk/raw/main/Serif/OTF/SimplifiedChinese/NotoSerifCJKsc-Regular.otf'
                simsun_path = os.path.join(fonts_dir, 'NotoSerifCJKsc-Regular.otf')
                if not os.path.exists(simsun_path):
                    urllib.request.urlretrieve(url, simsun_path)
            except Exception:
                simsun_path = None

    if simsun_path and os.path.exists(simsun_path):
        fm.fontManager.addfont(simsun_path)

    # --- 同时加载已有的 simhei.ttf 作为备用 ---
    simhei_path = os.path.join(fonts_dir, 'simhei.ttf')
    if os.path.exists(simhei_path):
        fm.fontManager.addfont(simhei_path)

    # --- Times New Roman ---
    tnr_path = os.path.join(fonts_dir, 'times.ttf')
    if not os.path.exists(tnr_path):
        sys_tnr = [
            '/usr/share/fonts/truetype/msttcorefonts/Times_New_Roman.ttf',
            '/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf',
        ]
        copied = False
        for p in sys_tnr:
            if os.path.exists(p):
                import shutil
                shutil.copy2(p, tnr_path)
                copied = True
                break
        if not copied:
            try:
                url = 'https://github.com/matomo-org/travis-scripts/raw/master/fonts/Times_New_Roman.ttf'
                if not os.path.exists(tnr_path):
                    urllib.request.urlretrieve(url, tnr_path)
            except Exception:
                tnr_path = None

    if tnr_path and os.path.exists(tnr_path):
        fm.fontManager.addfont(tnr_path)

    available = {f.name for f in fm.fontManager.ttflist}

    # 中文优先宋体
    chinese_candidates = ['NotoSerifCJKsc-Regular', 'Noto Serif CJK SC', 'SimSun', 'NSimSun',
                          'STSong', 'SimHei', 'WenQuanYi Zen Hei', 'FangSong']
    # 英文优先 Times New Roman
    english_candidates = ['Times New Roman', 'Times', 'Liberation Serif', 'DejaVu Serif']

    chinese_font = next((f for f in chinese_candidates if f in available), None)
    english_font = next((f for f in english_candidates if f in available), 'DejaVu Serif')

    return chinese_font, english_font


def make_bar_chart(config):
    """
    生成分组柱状图（类OriginPro风格）
    config keys:
      x_col, y_cols (list), std_cols (list or []), label_cols (list or []),
      colors (list), x_label, y_label, title, data (list of dicts)
    """
    chinese_font, english_font = setup_fonts()

    data = config['data']
    x_col = config['x_col']
    y_cols = config['y_cols']
    std_cols = config.get('std_cols', [])
    label_cols = config.get('label_cols', [])
    colors = config.get('colors', [])
    x_label = config.get('x_label', '')
    y_label = config.get('y_label', '')
    title = config.get('title', '')

    # 字号设置（前端可传入，否则用默认值）
    fs = config.get('font_sizes', {})
    fs_title      = int(fs.get('title', 14))
    fs_axis_label = int(fs.get('axis_label', 13))
    fs_tick       = int(fs.get('tick', 11))
    fs_legend     = int(fs.get('legend', 10))
    fs_data_label = int(fs.get('data_label', 9))

    bold = config.get('bold', False)
    if isinstance(bold, dict):
        fw_title = 'bold' if bold.get('title') else 'normal'
        fw_axis = 'bold' if bold.get('axis_label') else 'normal'
        fw_tick = 'bold' if bold.get('tick') else 'normal'
        fw_legend = 'bold' if bold.get('legend') else 'normal'
        fw_data = 'bold' if bold.get('data_label') else 'normal'
    else:
        fw_title = fw_axis = fw_tick = fw_legend = fw_data = 'bold' if bold else 'normal'

    df = pd.DataFrame(data)
    x_vals = df[x_col].tolist()
    n_groups = len(x_vals)
    n_series = len(y_cols)
    default_colors = [
        '#8FBC8F',  # 绿
        '#DAA520',  # 金黄
        '#87CEEB',  # 蓝
        '#E8A0A0',  # 粉红
        '#9B8EC4',  # 紫
        '#F4A460',  # 棕
    ]
    bar_colors = []
    for i in range(n_series):
        if i < len(colors) and colors[i]:
            bar_colors.append(colors[i])
        else:
            bar_colors.append(default_colors[i % len(default_colors)])

    # 图形尺寸（高分辨率）
    fig_w = max(8, n_groups * n_series * 0.6 + 3)
    fig, ax = plt.subplots(figsize=(fig_w, 6), dpi=150)

    group_spacing = float(config.get('bar_group_spacing', 0.3))
    bar_width = (1.0 - group_spacing) / n_series if n_series > 0 else 0.7
    bar_width = min(bar_width, 0.9 / n_series)
    bar_fill = float(config.get('bar_inner_gap', 0.9))  # 组内填充比，越小间隙越大
    x_pos = np.arange(n_groups)

    for si, ycol in enumerate(y_cols):
        y_vals = pd.to_numeric(df[ycol], errors='coerce').tolist()
        offset = (si - (n_series - 1) / 2) * bar_width

        # 误差棒
        yerr = None
        if si < len(std_cols) and std_cols[si] and std_cols[si] in df.columns:
            yerr = pd.to_numeric(df[std_cols[si]], errors='coerce').tolist()

        bars = ax.bar(
            x_pos + offset, y_vals,
            width=bar_width * bar_fill,
            color=bar_colors[si],
            label=strip_stat_suffix(ycol),
            edgecolor='black',
            linewidth=0.8,
            zorder=3
        )

        if yerr is not None:
            ax.errorbar(
                x_pos + offset, y_vals,
                yerr=yerr,
                fmt='none',
                ecolor='black',
                elinewidth=1,
                capsize=4,
                capthick=1,
                zorder=4
            )

        # Duncan / 标签
        if si < len(label_cols) and label_cols[si] and label_cols[si] in df.columns:
            labels = df[label_cols[si]].tolist()
            for xi, (bar, lbl, yv) in enumerate(zip(bars, labels, y_vals)):
                if lbl and str(lbl).strip():
                    top = yv + (yerr[xi] if yerr else 0)
                    font_kw = {}
                    if chinese_font:
                        font_kw['fontfamily'] = chinese_font
                    ax.text(
                        bar.get_x() + bar.get_width() / 2,
                        top + ax.get_ylim()[1] * 0.01,
                        str(lbl),
                        ha='center', va='bottom',
                        fontsize=fs_data_label,
                        **font_kw
                    )

    # 坐标轴样式（类OriginPro）
    ax.set_xticks(x_pos)
    ax.set_xticklabels(x_vals, fontsize=fs_tick)
    ax.set_xlim(-0.6, n_groups - 0.4)
    ax.set_ylim(bottom=0)

    # 字体设置
    tick_font = {}
    label_font = {}
    if english_font:
        tick_font['fontfamily'] = english_font
        label_font['fontfamily'] = english_font
    for tick in ax.get_xticklabels() + ax.get_yticklabels():
        tick.set_fontsize(fs_tick)
        tick.set_fontweight(fw_tick)
        if english_font:
            tick.set_fontfamily(english_font)

    if x_label:
        kw = {'fontsize': fs_axis_label, 'fontweight': fw_axis}
        if chinese_font and any('\u4e00' <= c <= '\u9fff' for c in x_label):
            kw['fontfamily'] = chinese_font
        elif english_font:
            kw['fontfamily'] = english_font
        ax.set_xlabel(x_label, **kw)

    if y_label:
        kw = {'fontsize': fs_axis_label, 'fontweight': fw_axis}
        if chinese_font and any('\u4e00' <= c <= '\u9fff' for c in y_label):
            kw['fontfamily'] = chinese_font
        elif english_font:
            kw['fontfamily'] = english_font
        ax.set_ylabel(y_label, **kw)

    if title:
        kw = {'fontsize': fs_title, 'fontweight': fw_title}
        if chinese_font and any('\u4e00' <= c <= '\u9fff' for c in title):
            kw['fontfamily'] = chinese_font
        elif english_font:
            kw['fontfamily'] = english_font
        ax.set_title(title, **kw)

    # 图例（单系列也显示，方便识别）
    legend_kw = {'fontsize': fs_legend, 'frameon': True, 'edgecolor': '#cccccc'}
    if chinese_font:
        legend_kw['prop'] = fm.FontProperties(family=chinese_font, size=fs_legend, weight=fw_legend)
    ax.legend(**legend_kw)

    # 去掉上/右边框（OriginPro风格）
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_linewidth(1.2)
    ax.spines['bottom'].set_linewidth(1.2)

    # 网格线控制
    show_grid = config.get('show_grid', False)
    if show_grid:
        ax.yaxis.grid(True, linestyle='--', alpha=0.4, zorder=0)
        ax.set_axisbelow(True)

    plt.tight_layout()

    # 输出 PNG (base64)
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    img_b64 = base64.b64encode(buf.read()).decode('utf-8')
    plt.close(fig)

    # 同时生成可编辑的 Excel（数据表）
    excel_buf = io.BytesIO()
    out_df = df[[x_col] + y_cols +
                [c for c in std_cols if c and c in df.columns] +
                [c for c in label_cols if c and c in df.columns]]
    out_df.to_excel(excel_buf, index=False)
    excel_buf.seek(0)
    excel_b64 = base64.b64encode(excel_buf.read()).decode('utf-8')

    return img_b64, excel_b64


def make_line_chart(config):
    """
    生成折线图（类OriginPro风格）
    config keys:
      x_col, y_cols (list), std_cols (list or []), label_cols (list or []),
      colors (list), x_label, y_label, title, data (list of dicts),
      line_styles (list), line_widths (list)
    """
    chinese_font, english_font = setup_fonts()

    data = config['data']
    x_col = config['x_col']
    y_cols = config['y_cols']
    std_cols = config.get('std_cols', [])
    label_cols = config.get('label_cols', [])
    colors = config.get('colors', [])
    line_styles = config.get('line_styles', [])  # 线条样式：'-', '--', '-.', ':'
    line_widths = config.get('line_widths', [])  # 线条粗细
    x_label = config.get('x_label', '')
    y_label = config.get('y_label', '')
    title = config.get('title', '')

    # 字号设置
    fs = config.get('font_sizes', {})
    fs_title      = int(fs.get('title', 14))
    fs_axis_label = int(fs.get('axis_label', 13))
    fs_tick       = int(fs.get('tick', 11))
    fs_legend     = int(fs.get('legend', 10))
    fs_data_label = int(fs.get('data_label', 9))

    bold = config.get('bold', False)
    if isinstance(bold, dict):
        fw_title = 'bold' if bold.get('title') else 'normal'
        fw_axis = 'bold' if bold.get('axis_label') else 'normal'
        fw_tick = 'bold' if bold.get('tick') else 'normal'
        fw_legend = 'bold' if bold.get('legend') else 'normal'
        fw_data = 'bold' if bold.get('data_label') else 'normal'
    else:
        fw_title = fw_axis = fw_tick = fw_legend = fw_data = 'bold' if bold else 'normal'

    df = pd.DataFrame(data)
    x_vals = df[x_col].tolist()
    n_groups = len(x_vals)
    n_series = len(y_cols)

    # 默认颜色
    default_colors = [
        '#000000',  # 黑色
        '#FF0000',  # 红色
        '#0000FF',  # 蓝色
        '#008000',  # 绿色
        '#FF00FF',  # 紫色
        '#FFA500',  # 橙色
    ]
    line_colors = []
    for i in range(n_series):
        if i < len(colors) and colors[i]:
            line_colors.append(colors[i])
        else:
            line_colors.append(default_colors[i % len(default_colors)])

    # 默认线条样式
    default_line_styles = ['-', '-', '-', '-', '-', '-']
    line_style_list = []
    for i in range(n_series):
        if i < len(line_styles) and line_styles[i]:
            line_style_list.append(line_styles[i])
        else:
            line_style_list.append(default_line_styles[i % len(default_line_styles)])

    # 默认线条粗细
    default_line_widths = [2, 2, 2, 2, 2, 2]
    line_width_list = []
    for i in range(n_series):
        if i < len(line_widths) and line_widths[i]:
            line_width_list.append(float(line_widths[i]))
        else:
            line_width_list.append(default_line_widths[i % len(default_line_widths)])

    # 图形尺寸
    fig_w = max(8, n_groups * 0.8 + 3)
    fig, ax = plt.subplots(figsize=(fig_w, 6), dpi=150)

    # 转换 x 轴为数值（如果可能）
    try:
        x_numeric = pd.to_numeric(df[x_col], errors='coerce').tolist()
        if all(pd.notna(x_numeric)):
            x_positions = x_numeric
        else:
            x_positions = list(range(n_groups))
    except:
        x_positions = list(range(n_groups))

    for si, ycol in enumerate(y_cols):
        y_vals = pd.to_numeric(df[ycol], errors='coerce').tolist()

        # 误差棒
        yerr = None
        if si < len(std_cols) and std_cols[si] and std_cols[si] in df.columns:
            yerr = pd.to_numeric(df[std_cols[si]], errors='coerce').tolist()

        # 绘制折线
        line = ax.plot(
            x_positions, y_vals,
            color=line_colors[si],
            linestyle=line_style_list[si],
            linewidth=line_width_list[si],
            marker='o',
            markersize=6,
            label=strip_stat_suffix(ycol),
            zorder=3
        )[0]

        # 误差棒
        if yerr is not None:
            ax.errorbar(
                x_positions, y_vals,
                yerr=yerr,
                fmt='none',
                ecolor=line_colors[si],
                elinewidth=1,
                capsize=4,
                capthick=1,
                zorder=2
            )

        # Duncan / 标签
        if si < len(label_cols) and label_cols[si] and label_cols[si] in df.columns:
            labels = df[label_cols[si]].tolist()
            for xi, (xp, lbl, yv) in enumerate(zip(x_positions, labels, y_vals)):
                if lbl and str(lbl).strip():
                    top = yv + (yerr[xi] if yerr else 0)
                    font_kw = {}
                    if chinese_font:
                        font_kw['fontfamily'] = chinese_font
                    ax.text(
                        xp,
                        top + (ax.get_ylim()[1] - ax.get_ylim()[0]) * 0.02,
                        str(lbl),
                        ha='center', va='bottom',
                        fontsize=fs_data_label,
                        color=line_colors[si],
                        **font_kw
                    )

    # 坐标轴样式
    ax.set_xticks(x_positions)
    ax.set_xticklabels(x_vals, fontsize=fs_tick)

    # 字体设置
    tick_font = {}
    label_font = {}
    if english_font:
        tick_font['fontfamily'] = english_font
        label_font['fontfamily'] = english_font
    for tick in ax.get_xticklabels() + ax.get_yticklabels():
        tick.set_fontsize(fs_tick)
        tick.set_fontweight(fw_tick)
        if english_font:
            tick.set_fontfamily(english_font)

    if x_label:
        kw = {'fontsize': fs_axis_label, 'fontweight': fw_axis}
        if chinese_font and any('\u4e00' <= c <= '\u9fff' for c in x_label):
            kw['fontfamily'] = chinese_font
        elif english_font:
            kw['fontfamily'] = english_font
        ax.set_xlabel(x_label, **kw)

    if y_label:
        kw = {'fontsize': fs_axis_label, 'fontweight': fw_axis}
        if chinese_font and any('\u4e00' <= c <= '\u9fff' for c in y_label):
            kw['fontfamily'] = chinese_font
        elif english_font:
            kw['fontfamily'] = english_font
        ax.set_ylabel(y_label, **kw)

    if title:
        kw = {'fontsize': fs_title, 'fontweight': fw_title}
        if chinese_font and any('\u4e00' <= c <= '\u9fff' for c in title):
            kw['fontfamily'] = chinese_font
        elif english_font:
            kw['fontfamily'] = english_font
        ax.set_title(title, **kw)

    # 图例
    legend_kw = {'fontsize': fs_legend, 'frameon': True, 'edgecolor': '#cccccc'}
    if chinese_font:
        legend_kw['prop'] = fm.FontProperties(family=chinese_font, size=fs_legend, weight=fw_legend)
    ax.legend(**legend_kw)

    # 去掉上/右边框
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_linewidth(1.2)
    ax.spines['bottom'].set_linewidth(1.2)

    # 网格线控制
    show_grid = config.get('show_grid', False)
    if show_grid:
        ax.yaxis.grid(True, linestyle='--', alpha=0.4, zorder=0)
        ax.set_axisbelow(True)

    plt.tight_layout()

    # 输出 PNG (base64)
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    img_b64 = base64.b64encode(buf.read()).decode('utf-8')
    plt.close(fig)

    # 同时生成可编辑的 Excel（数据表）
    excel_buf = io.BytesIO()
    out_df = df[[x_col] + y_cols +
                [c for c in std_cols if c and c in df.columns] +
                [c for c in label_cols if c and c in df.columns]]
    out_df.to_excel(excel_buf, index=False)
    excel_buf.seek(0)
    excel_b64 = base64.b64encode(excel_buf.read()).decode('utf-8')

    return img_b64, excel_b64




def make_heatmap(config):
    """
    生成热图（聚类热图风格，参考图1）
    config keys: data, row_col, value_cols, normalize, title, font_sizes, bold
    """
    from scipy.cluster.hierarchy import linkage, dendrogram
    from scipy.spatial.distance import pdist
    chinese_font, english_font = setup_fonts()

    data = config['data']
    row_col = config.get('row_col')
    value_cols = config.get('value_cols', [])
    normalize = config.get('normalize', True)
    title = config.get('title', '')
    fs = config.get('font_sizes', {})
    fs_title = int(fs.get('title', 12))
    fs_row = int(fs.get('tick', 8))
    fs_col = int(fs.get('axis_label', 10))
    bold = config.get('bold', False)
    cluster_rows = config.get('cluster_rows', True)
    cluster_cols = config.get('cluster_cols', True)
    # 图例色阶参数
    vmin_cfg = config.get('vmin', None)
    vmax_cfg = config.get('vmax', None)
    vstep_cfg = float(config.get('vstep', 0.5))
    color_min = config.get('color_min', '#1446AF')
    color_mid = config.get('color_mid', '#EDE7AD')
    color_max = config.get('color_max', '#CF1C1D')
    cbar_fontsize = int(config.get('cbar_fontsize', 7))
    cbar_bold = config.get('cbar_bold', False)
    cbar_fw = 'bold' if cbar_bold else 'normal'
    if isinstance(bold, dict):
        fw_tick = 'bold' if bold.get('tick') else 'normal'
        fw_axis = 'bold' if bold.get('axis_label') else 'normal'
    else:
        fw_tick = fw_axis = 'bold' if bold else 'normal'
    fw = fw_tick  # backward compat alias

    df = pd.DataFrame(data)
    if row_col and row_col in df.columns:
        df = df.set_index(row_col)
    # 只保留数值列
    if value_cols:
        df = df[value_cols]

    # 强制转换为数值类型，处理所有非数值数据
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    df = df.fillna(0)

    # 确保所有数据都是float类型
    df = df.astype(float)

    # 检查是否有有效数据
    if df.empty or df.shape[0] == 0 or df.shape[1] == 0:
        raise ValueError("没有有效的数值数据")

    # 归一化（行z-score）
    if normalize:
        from scipy.stats import zscore
        mat = pd.DataFrame(index=df.index, columns=df.columns)
        for idx in df.index:
            row_data = df.loc[idx].values
            if np.std(row_data) > 0:
                mat.loc[idx] = zscore(row_data)
            else:
                mat.loc[idx] = row_data - np.mean(row_data)
    else:
        mat = df.copy()

    # 确保mat_arr是纯数值的numpy数组
    mat_arr = np.array(mat.values, dtype=float)
    n_rows, n_cols = mat_arr.shape

    # TBtools标准聚类：ward linkage + correlation距离
    row_link = None
    if cluster_rows and n_rows > 1:
        try:
            row_link = linkage(pdist(mat_arr, metric='correlation'), method='ward')
        except Exception:
            row_link = linkage(pdist(mat_arr, metric='euclidean'), method='ward')
        row_dend = dendrogram(row_link, no_plot=True)
        row_order = row_dend['leaves']
    else:
        row_order = list(range(n_rows))

    col_link = None
    if cluster_cols and n_cols > 1:
        try:
            col_link = linkage(pdist(mat_arr.T, metric='correlation'), method='ward')
        except Exception:
            col_link = linkage(pdist(mat_arr.T, metric='euclidean'), method='ward')
        col_dend = dendrogram(col_link, no_plot=True)
        col_order = col_dend['leaves']
    else:
        col_order = list(range(n_cols))

    mat_ordered = mat_arr[np.ix_(row_order, col_order)]
    row_labels = [mat.index[i] for i in row_order]
    col_labels = [mat.columns[i] for i in col_order]

    # 字体：优先宋体，备用其他中文字体
    from matplotlib import font_manager
    zh_candidates = ['NotoSerifCJKsc-Regular', 'Noto Serif CJK SC', 'SimSun', 'NSimSun',
                     'STSong', 'SimHei', 'Microsoft YaHei', 'WenQuanYi Micro Hei',
                     'Noto Sans CJK SC', 'AR PL UMing CN', 'Source Han Sans CN', 'PingFang SC']
    available_names = {f.name for f in font_manager.fontManager.ttflist}
    zh_font = next((f for f in zh_candidates if f in available_names), chinese_font)

    # 全局设置中文字体，确保所有标签都能显示中文
    if zh_font:
        plt.rcParams['font.family'] = zh_font
        plt.rcParams['axes.unicode_minus'] = False

    def apply_font(labels):
        for lbl in labels:
            text = lbl.get_text()
            if zh_font and any('\u4e00' <= c <= '\u9fff' for c in text):
                lbl.set_fontfamily(zh_font)
            elif english_font:
                lbl.set_fontfamily(english_font)

    # 图形尺寸：根据行数和标签长度动态调整
    row_label_maxlen = max((len(str(l)) for l in row_labels), default=1)
    # 中文字符宽度约为英文的2倍，估算标签宽度
    label_inch = max(2.5, row_label_maxlen * 0.18)
    fig_h = max(8, n_rows * 0.32 + 3)
    fig_w = max(12, n_cols * 1.5 + label_inch + 5)
    fig = plt.figure(figsize=(fig_w, fig_h), dpi=150)

    # 布局比例：左侧行树状图 | 热图 | 右侧行标签 + 色条
    from matplotlib.gridspec import GridSpec
    # 根据标签长度动态计算右侧留白
    right_margin = min(0.75, max(0.55, 1.0 - label_inch / fig_w - 0.08))
    gs = GridSpec(2, 4,
                  width_ratios=[0.10, 0.50, 0.02, 0.04],
                  height_ratios=[0.08, 0.92],
                  left=0.02, right=right_margin, top=0.95, bottom=0.06,
                  hspace=0.01, wspace=0.02)

    ax_col_dend = fig.add_subplot(gs[0, 1])
    ax_hmap = fig.add_subplot(gs[1, 1])
    ax_row_dend = fig.add_subplot(gs[1, 0])

    # 颜色映射：使用用户自定义颜色
    from matplotlib.colors import LinearSegmentedColormap
    cmap = LinearSegmentedColormap.from_list('custom',
        [color_min, color_mid, color_max])
    # vmin/vmax：优先用户设置，否则自动计算
    if vmin_cfg is not None and vmax_cfg is not None:
        vmin, vmax = float(vmin_cfg), float(vmax_cfg)
    else:
        vmax = np.percentile(np.abs(mat_ordered), 98) or 1
        vmin = -vmax

    # 色块：用pcolormesh，坐标系0~n_rows/n_cols
    im = ax_hmap.pcolormesh(mat_ordered, cmap=cmap, vmin=vmin, vmax=vmax,
                            edgecolors='white', linewidth=0.5)
    ax_hmap.set_xlim(0, n_cols)
    ax_hmap.set_ylim(0, n_rows)
    ax_hmap.invert_yaxis()

    # 行标签（右侧）：行数多时自动缩小字号
    auto_fs_row = min(fs_row, max(5, int(fig_h * 72 * 0.85 / n_rows * 0.7))) if n_rows > 20 else fs_row
    ax_hmap.set_yticks([i + 0.5 for i in range(n_rows)])
    ax_hmap.set_yticklabels(row_labels, fontsize=auto_fs_row, fontweight=fw_tick)
    apply_font(ax_hmap.get_yticklabels())
    ax_hmap.yaxis.set_tick_params(length=0, pad=4)
    ax_hmap.yaxis.set_label_position('right')
    ax_hmap.yaxis.tick_right()

    # 列标签（底部）
    ax_hmap.set_xticks([i + 0.5 for i in range(n_cols)])
    ax_hmap.set_xticklabels(col_labels, fontsize=fs_col, fontweight=fw_axis, rotation=0)
    apply_font(ax_hmap.get_xticklabels())
    ax_hmap.xaxis.set_tick_params(length=0, pad=4)
    ax_hmap.xaxis.set_label_position('bottom')

    # 行树状图：精确对齐
    # pcolormesh y轴: invert后上=0, 下=n_rows; dendrogram leaf i 坐标=(i+0.5)*10
    # 所以 ylim 需要 top=0, bottom=n_rows*10 → set_ylim(n_rows*10, 0)
    if row_link is not None:
        dendrogram(row_link, ax=ax_row_dend, orientation='left',
                   labels=mat.index.tolist(), no_labels=True,
                   link_color_func=lambda k: 'black',
                   above_threshold_color='black')
        ax_row_dend.set_ylim(n_rows * 10, 0)
    ax_row_dend.axis('off')

    # 列树状图
    if col_link is not None:
        col_dend_data = dendrogram(col_link, ax=ax_col_dend, orientation='top',
                   labels=mat.columns.tolist(), no_labels=True,
                   link_color_func=lambda k: 'black',
                   above_threshold_color='black')
        ax_col_dend.set_xlim(0, n_cols * 10)
    ax_col_dend.axis('off')

    # 色条：小尺寸放右上角，刻度按vstep设置
    cbar_ax = fig.add_axes([0.91, 0.70, 0.025, 0.22])
    cbar = fig.colorbar(im, cax=cbar_ax)
    import numpy as np_
    tick_vals = np_.arange(vmin, vmax + vstep_cfg * 0.01, vstep_cfg)
    cbar.set_ticks(tick_vals)
    cbar.ax.tick_params(labelsize=cbar_fontsize)
    for lbl in cbar.ax.get_yticklabels():
        lbl.set_fontweight(cbar_fw)
        if english_font:
            lbl.set_fontfamily(english_font)

    # 标题
    if title:
        kw = {'fontsize': fs_title, 'fontweight': fw}
        if zh_font and any('\u4e00' <= c <= '\u9fff' for c in title):
            kw['fontfamily'] = zh_font
        fig.suptitle(title, **kw)

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor='white')
    buf.seek(0)
    img_b64 = base64.b64encode(buf.read()).decode('utf-8')
    plt.close(fig)

    # Excel 输出
    excel_buf = io.BytesIO()
    out_df = pd.DataFrame(mat_ordered, index=row_labels, columns=col_labels)
    out_df.to_excel(excel_buf, index=True)
    excel_buf.seek(0)
    excel_b64 = base64.b64encode(excel_buf.read()).decode('utf-8')

    return img_b64, excel_b64


def make_radar_chart(config):
    """
    生成雷达图
    config keys:
      data (list of dicts), axes_col (轴标签列), series_cols (list of series columns),
      colors (list), line_styles (list), line_widths (list), marker_styles (list),
      title, font_sizes, transpose (bool)
    """
    chinese_font, english_font = setup_fonts()

    data = config['data']
    axes_col = config['axes_col']  # 雷达图的轴标签列
    series_cols = config['series_cols']  # 每个系列对应一列数据
    colors = config.get('colors', [])
    line_styles = config.get('line_styles', [])
    line_widths = config.get('line_widths', [])
    marker_styles = config.get('marker_styles', [])
    title = config.get('title', '')
    transpose = config.get('transpose', False)  # 是否转置数据

    # 字号设置
    fs = config.get('font_sizes', {})
    fs_title = int(fs.get('title', 14))
    fs_axis_label = int(fs.get('axis_label', 11))
    fs_tick = int(fs.get('tick', 10))
    fs_legend = int(fs.get('legend', 10))

    df = pd.DataFrame(data)

    # 如果需要转置
    if transpose:
        # 原始数据：行为样品，列为指标
        # 转置后：行为指标，列为样品
        # axes_col 变成第一列（指标名称）
        # series_cols 变成行（样品名称）

        # 提取需要的列
        transpose_cols = [axes_col] + series_cols
        df_subset = df[transpose_cols].copy()

        # 转置：第一列变成列名，其他列变成行
        df_transposed = df_subset.set_index(axes_col).T
        df_transposed.reset_index(inplace=True)
        df_transposed.rename(columns={'index': 'sample'}, inplace=True)

        # 更新变量
        df = df_transposed
        axes_col = df.columns[1]  # 第一个指标列作为轴标签列
        series_cols = df['sample'].tolist()  # 样品名称作为系列

        # 重新构建数据：每行是一个指标，每列是一个样品
        axes_labels = df.columns[1:].tolist()
        n_axes = len(axes_labels)
    else:
        # 标准格式：行为指标，列为样品
        axes_labels = df[axes_col].tolist()
        n_axes = len(axes_labels)

    # 计算角度（从顶部开始，顺时针）
    angles = np.linspace(0, 2 * np.pi, n_axes, endpoint=False).tolist()
    angles += angles[:1]  # 闭合

    # 提前计算Y轴范围（绘制时需要clip）
    import math
    y_min = config.get('y_min', None)
    y_max = config.get('y_max', None)
    if y_min is None or y_max is None:
        all_vals = []
        if transpose:
            for sample_name in series_cols:
                sample_row = df[df['sample'] == sample_name]
                if not sample_row.empty:
                    for axis_label in axes_labels:
                        if axis_label in sample_row.columns:
                            v = pd.to_numeric(sample_row[axis_label].iloc[0], errors='coerce')
                            if pd.notna(v):
                                all_vals.append(v)
        else:
            for sc in series_cols:
                if sc in df.columns:
                    all_vals.extend(pd.to_numeric(df[sc], errors='coerce').dropna().tolist())
        if all_vals:
            data_min, data_max = min(all_vals), max(all_vals)
            span = data_max - data_min if data_max != data_min else 1
            if y_min is None:
                y_min = data_min - span * 0.05
            if y_max is None:
                y_max = data_max + span * 0.1
        else:
            y_min, y_max = 0, 1
    y_min, y_max = float(y_min), float(y_max)

    # 默认颜色
    default_colors = ['#000000', '#FF0000', '#00FF00', '#0000FF', '#FF00FF', '#FFA500', '#00FFFF']
    # 默认线条样式
    default_line_styles = ['-', '-', '-', '-', '-', '-', '-']
    # 默认线条粗细
    default_line_widths = [2, 2, 2, 2, 2, 2, 2]
    # 默认标记样式
    default_markers = ['o', 's', '^', 'D', 'v', '<', '>']

    # 创建图形
    fig = plt.figure(figsize=(8, 8), dpi=150)
    ax = fig.add_subplot(111, projection='polar')
    ax.set_theta_zero_location('N')   # 0度在顶部
    ax.set_theta_direction(-1)        # 顺时针排列（与图1一致）

    # 绘制每个系列
    if transpose:
        # 转置模式：series_cols 是样品名称，需要从 df 中找到对应的行
        for si, sample_name in enumerate(series_cols):
            # 找到该样品的行
            sample_row = df[df['sample'] == sample_name]
            if sample_row.empty:
                continue

            # 提取所有指标列的值
            values = []
            for axis_label in axes_labels:
                if axis_label in sample_row.columns:
                    val = pd.to_numeric(sample_row[axis_label].iloc[0], errors='coerce')
                    values.append(val if pd.notna(val) else 0)
                else:
                    values.append(0)
            values += values[:1]  # 闭合
            # 裁剪到y轴范围（此处先存储，后面统一裁剪）
            values = [np.clip(v, y_min, y_max) for v in values]

            # 获取颜色
            color = colors[si] if si < len(colors) and colors[si] else default_colors[si % len(default_colors)]
            # 获取线条样式
            linestyle = line_styles[si] if si < len(line_styles) and line_styles[si] else default_line_styles[si % len(default_line_styles)]
            # 获取线条粗细
            linewidth = float(line_widths[si]) if si < len(line_widths) and line_widths[si] else default_line_widths[si % len(default_line_widths)]
            # 获取标记样式
            marker = marker_styles[si] if si < len(marker_styles) and marker_styles[si] else default_markers[si % len(default_markers)]

            ax.plot(angles, values,
                    color=color,
                    linestyle=linestyle,
                    linewidth=linewidth,
                    marker=marker,
                    markersize=6,
                    label=strip_stat_suffix(sample_name))
            ax.fill(angles, values, color=color, alpha=0.1)
    else:
        # 标准模式：series_cols 是列名
        for si, series_col in enumerate(series_cols):
            if series_col not in df.columns:
                continue

            values = pd.to_numeric(df[series_col], errors='coerce').tolist()
            values += values[:1]  # 闭合
            values = [np.clip(v, y_min, y_max) for v in values]

            # 获取颜色
            color = colors[si] if si < len(colors) and colors[si] else default_colors[si % len(default_colors)]
            # 获取线条样式
            linestyle = line_styles[si] if si < len(line_styles) and line_styles[si] else default_line_styles[si % len(default_line_styles)]
            # 获取线条粗细
            linewidth = float(line_widths[si]) if si < len(line_widths) and line_widths[si] else default_line_widths[si % len(default_line_widths)]
            # 获取标记样式
            marker = marker_styles[si] if si < len(marker_styles) and marker_styles[si] else default_markers[si % len(default_markers)]

            ax.plot(angles, values,
                    color=color,
                    linestyle=linestyle,
                    linewidth=linewidth,
                    marker=marker,
                    markersize=6,
                    label=strip_stat_suffix(series_col))
            ax.fill(angles, values, color=color, alpha=0.1)

    ax.set_ylim(y_min, y_max)

    # 生成刻度值
    y_step = config.get('y_step', None)
    if y_step and float(y_step) > 0:
        tick_step = float(y_step)
    else:
        tick_count = 5
        tick_step = (y_max - y_min) / tick_count
        magnitude = 10 ** math.floor(math.log10(abs(tick_step))) if tick_step > 0 else 1
        tick_step = round(tick_step / magnitude) * magnitude or magnitude
    # 包含起点的刻度列表
    tick_start = math.ceil(y_min / tick_step + 1e-9) * tick_step
    yticks = [round(y_min, 10)]  # 始终包含起点
    v = tick_start
    while v <= y_max + 1e-9:
        val = round(v, 10)
        if abs(val - y_min) > tick_step * 0.01:  # 避免重复
            yticks.append(val)
        v += tick_step
    ax.set_yticks(yticks)

    # 刻度标签位置偏移（避免与数据线重叠）
    label_angle_deg = float(config.get('tick_label_angle', 45))
    bold = config.get('bold', False)
    if isinstance(bold, dict):
        fw_tick = 'bold' if bold.get('tick') else 'normal'
        fw_axis = 'bold' if bold.get('axis_label') else 'normal'
        fw_legend = 'bold' if bold.get('legend') else 'normal'
    else:
        fw_tick = fw_axis = fw_legend = 'bold' if bold else 'normal'
    fw = fw_tick
    ax.set_rlabel_position(label_angle_deg)
    # 隐藏默认刻度标签，手动绘制以确保始终水平
    ax.set_yticklabels([])
    label_angle_rad = math.radians(label_angle_deg)
    for ytick in yticks:
        if ytick < y_min - 1e-9 or ytick > y_max + 1e-9:
            continue
        txt = f'{ytick:g}'
        kw_tick = dict(fontsize=fs_tick, ha='left', va='center', fontweight=fw,
                       rotation=0, zorder=5)
        if english_font:
            kw_tick['fontfamily'] = english_font
        ax.text(label_angle_rad, ytick, txt, **kw_tick)

    # 多边形网格（手动绘制）
    ax.yaxis.grid(False)
    ax.xaxis.grid(False)
    ax.spines['polar'].set_visible(False)

    # 指标标签与轴的距离
    axis_label_pad = float(config.get('axis_label_pad', 20))
    ax.tick_params(axis='x', pad=axis_label_pad)

    # 重新设置轴标签
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(axes_labels, fontsize=fs_axis_label, fontweight=fw_axis)
    if chinese_font:
        for label in ax.get_xticklabels():
            label.set_fontfamily(chinese_font)
            label.set_fontweight(fw_axis)
    elif english_font:
        for label in ax.get_xticklabels():
            label.set_fontfamily(english_font)

    # 绘制多边形背景网格
    grid_color = config.get('grid_color', 'gray')
    grid_width = float(config.get('grid_width', 0.8))
    spoke_color = config.get('spoke_color', 'gray')
    spoke_width = float(config.get('spoke_width', 1.2))
    for ytick in yticks:
        if ytick < y_min - 1e-9 or ytick > y_max + 1e-9:
            continue
        poly_r = [ytick] * (n_axes + 1)
        poly_a = angles[:-1] + [angles[0]]
        ax.plot(poly_a, poly_r, color=grid_color, linewidth=grid_width, linestyle='-', alpha=0.5, zorder=0)

    # 绘制从中心到各顶点的辐射线
    for angle in angles[:-1]:
        ax.plot([angle, angle], [y_min, y_max], color=spoke_color, linewidth=spoke_width, linestyle='-', alpha=0.7, zorder=0)

    # 标题
    if title:
        kw = {'fontsize': fs_title, 'fontweight': 'bold', 'pad': 20}
        if chinese_font and any('\u4e00' <= c <= '\u9fff' for c in title):
            kw['fontfamily'] = chinese_font
        elif english_font:
            kw['fontfamily'] = english_font
        ax.set_title(title, **kw)

    # 图例
    legend_kw = {'fontsize': fs_legend, 'loc': 'upper right', 'bbox_to_anchor': (1.3, 1.1)}
    if chinese_font:
        legend_kw['prop'] = fm.FontProperties(family=chinese_font, size=fs_legend)
    ax.legend(**legend_kw)

    plt.tight_layout()

    # 输出 PNG (base64)
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    img_b64 = base64.b64encode(buf.read()).decode('utf-8')
    plt.close(fig)

    # 生成 Excel
    excel_buf = io.BytesIO()
    if transpose:
        # 转置模式：输出转置后的数据
        out_df = df
    else:
        # 标准模式：输出原始数据
        out_df = df[[axes_col] + [c for c in series_cols if c in df.columns]]
    out_df.to_excel(excel_buf, index=False)
    excel_buf.seek(0)
    excel_b64 = base64.b64encode(excel_buf.read()).decode('utf-8')

    return img_b64, excel_b64


def make_excel_with_chart(df, x_col, y_cols, std_cols, label_cols, colors):
    """
    用 openpyxl 生成含嵌入柱状图的 .xlsx 文件（Origin 可直接打开编辑）
    返回 base64 字符串
    """
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.styles import PatternFill, Font as XLFont, Alignment

        # 构建导出列
        export_cols = [x_col] + list(y_cols)
        for c in std_cols:
            if c and c in df.columns and c not in export_cols:
                export_cols.append(c)
        for c in label_cols:
            if c and c in df.columns and c not in export_cols:
                export_cols.append(c)

        wb = Workbook()
        ws = wb.active
        ws.title = '数据'

        # 写表头（用去后缀的名称）
        for ci, col in enumerate(export_cols, 1):
            cell = ws.cell(row=1, column=ci, value=strip_stat_suffix(col))
            cell.font = XLFont(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 写数据
        out_df = df[export_cols]
        for ri, row_vals in enumerate(out_df.itertuples(index=False), 2):
            for ci, val in enumerate(row_vals, 1):
                ws.cell(row=ri, column=ci, value=val)

        n_rows = len(out_df) + 1  # 含表头

        # 创建柱状图
        chart = BarChart()
        chart.type = 'col'
        chart.grouping = 'clustered'
        chart.overlap = 0
        chart.width = 22
        chart.height = 14

        # X 轴分类
        cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
        chart.set_categories(cats)

        # 每个 y_col 一个系列
        for si, ycol in enumerate(y_cols):
            col_idx = export_cols.index(ycol) + 1
            data_ref = Reference(ws, min_col=col_idx, min_row=1, max_row=n_rows)
            from openpyxl.chart import Series as XLSeries
            series = XLSeries(data_ref, title_from_data=True)
            # 设置系列颜色
            if si < len(colors) and colors[si]:
                hex_color = colors[si].lstrip('#').upper()
                from openpyxl.drawing.fill import PatternFillProperties
                series.graphicalProperties.solidFill = hex_color
            chart.series.append(series)

        # 将图表放在数据下方
        ws.add_chart(chart, f'A{n_rows + 3}')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return base64.b64encode(buf.read()).decode('utf-8')

    except Exception:
        return None


@app.route('/api/plot', methods=['POST'])
def api_plot():
    try:
        payload = request.get_json()
        if not payload:
            return jsonify({'error': '无效请求'}), 400

        chart_type = payload.get('chart_type', 'bar')
        if chart_type == 'bar':
            img_b64, excel_b64 = make_bar_chart(payload)
            # 生成含图表的 Excel（Origin 可直接打开编辑）
            df_plot = pd.DataFrame(payload['data'])
            opju_b64 = make_excel_with_chart(
                df_plot,
                payload['x_col'],
                payload['y_cols'],
                payload.get('std_cols', []),
                payload.get('label_cols', []),
                payload.get('colors', [])
            )
        elif chart_type == 'line':
            img_b64, excel_b64 = make_line_chart(payload)
            # 折线图也生成 Excel 数据文件
            df_plot = pd.DataFrame(payload['data'])
            opju_b64 = make_excel_with_chart(
                df_plot,
                payload['x_col'],
                payload['y_cols'],
                payload.get('std_cols', []),
                payload.get('label_cols', []),
                payload.get('colors', [])
            )
        elif chart_type == 'radar':
            img_b64, excel_b64 = make_radar_chart(payload)
            # 雷达图生成简单的 Excel 数据文件
            opju_b64 = excel_b64
        elif chart_type == 'heatmap':
            img_b64, excel_b64 = make_heatmap(payload)
            opju_b64 = excel_b64
        elif chart_type == 'pca':
            img_b64, excel_b64 = make_pca_plot(payload)
            opju_b64 = excel_b64
        else:
            return jsonify({'error': f'暂不支持图表类型: {chart_type}'}), 400

        return jsonify({'image': img_b64, 'excel': excel_b64, 'opju': opju_b64})

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


def make_pca_plot(config):
    """SIMCA 14.1 风格 PCA 得分图，支持手动分组、自定义颜色、轴范围、字体"""
    from sklearn.preprocessing import StandardScaler
    from sklearn.decomposition import PCA
    from scipy.stats import f as f_dist

    data = config.get('data', [])
    value_cols = config.get('value_cols', [])
    sample_col = config.get('sample_col', '')
    # groups_map: [{name, color, indices}] — 手动分组
    groups_map = config.get('groups_map', [])
    pc_x = int(config.get('pc_x', 1))
    pc_y = int(config.get('pc_y', 2))
    show_ellipse = config.get('show_ellipse', True)
    show_labels = config.get('show_labels', True)
    label_mode = config.get('label_mode', 'auto')
    xlabel_custom = config.get('xlabel', '')
    ylabel_custom = config.get('ylabel', '')
    show_bg = config.get('show_bg', True)
    show_grid = config.get('show_grid', True)
    title = config.get('title', '')
    x_min = config.get('x_min', None)
    x_max = config.get('x_max', None)
    y_min = config.get('y_min', None)
    y_max = config.get('y_max', None)
    axis_fontsize = int(config.get('axis_fontsize', 13))
    axis_color = config.get('axis_color', '#000000')
    tick_fontsize = int(config.get('tick_fontsize', 11))
    tick_color = config.get('tick_color', '#000000')
    label_fontsize = int(config.get('label_fontsize', 8))
    legend_fontsize = int(config.get('legend_fontsize', 10))
    dot_size = int(config.get('dot_size', 60))

    df = pd.DataFrame(data)
    df_num = df.apply(pd.to_numeric, errors='coerce')

    # 确定数值列
    if not value_cols:
        value_cols = [c for c in df_num.columns if df_num[c].notna().sum() >= 2]
    else:
        value_cols = [c for c in value_cols if c in df_num.columns and df_num[c].notna().sum() >= 2]

    if not value_cols:
        raise ValueError('没有找到有效的数值列，请检查数据或手动选择数值列')

    # 每行是一个样本（指标），每列是一个特征（处理组）
    X = df_num[value_cols].copy().fillna(df_num[value_cols].mean()).dropna()
    # 样本名取第一个非数值列（指标名列，如 Primary ID）
    label_col = sample_col if (sample_col and sample_col in df.columns) else df.columns[0]
    sample_names = df.loc[X.index, label_col].astype(str).tolist()

    if len(X) < 2:
        raise ValueError('有效数据行数不足（至少需要2行），请检查所选列是否包含数值数据')

    valid_idx = list(range(len(X)))
    n_components = max(pc_x, pc_y)
    X_scaled = StandardScaler().fit_transform(X)
    pca = PCA(n_components=n_components)
    scores = pca.fit_transform(X_scaled)
    r2x = pca.explained_variance_ratio_

    chinese_font, english_font = setup_fonts()
    efont = english_font or 'DejaVu Serif'

    # 构建分组信息
    # groups_map 格式: [{"name":"A","color":"#ff0000","indices":[0,1,2]}, ...]
    # indices 是 df 的行号（0-based in valid_idx）
    if not groups_map:
        groups_map = [{'name': 'All', 'color': '#1f77b4', 'indices': list(range(len(valid_idx)))}]

    # SIMCA 风格图形：宽扁比例
    fig, ax = plt.subplots(figsize=(12, 6), dpi=150)
    fig.patch.set_facecolor('white')
    ax.set_facecolor('#f0f0f0' if show_bg else 'white')

    for spine in ax.spines.values():
        spine.set_linewidth(0.8)
        spine.set_color('#888888')

    t_x = scores[:, pc_x - 1]
    t_y = scores[:, pc_y - 1]

    def draw_hotelling_ellipse(pts, color, lw=1.5, ls='-'):
        if len(pts) < 3:
            return
        mean = pts.mean(axis=0)
        cov = np.cov(pts.T)
        n, p = len(pts), 2
        if n <= p:
            return
        f_crit = f_dist.ppf(0.95, p, n - p)
        c2 = p * (n - 1) / (n - p) * f_crit
        try:
            vals, vecs = np.linalg.eigh(cov)
            vals = np.maximum(vals, 0)
            order = vals.argsort()[::-1]
            vals, vecs = vals[order], vecs[:, order]
            theta = np.linspace(0, 2 * np.pi, 300)
            ell = np.sqrt(c2) * (vecs * np.sqrt(vals)) @ np.array([np.cos(theta), np.sin(theta)])
            ax.plot(mean[0] + ell[0], mean[1] + ell[1], color=color, linewidth=lw, linestyle=ls, zorder=2)
        except Exception:
            pass

    # 全局椭圆（灰色实线，SIMCA风格）
    if show_ellipse:
        draw_hotelling_ellipse(np.column_stack([t_x, t_y]), '#888888', lw=1.8, ls='-')

    _all_texts = []
    _pt_x, _pt_y = [], []
    _label_data = []  # (x, y, label, color) for deferred placement
    for grp in groups_map:
        idxs = [i for i in grp.get('indices', []) if i < len(valid_idx)]
        if not idxs:
            continue
        color = grp.get('color', '#1f77b4')
        name = grp.get('name', '') or f'组{groups_map.index(grp)+1}'
        gx = t_x[idxs]
        gy = t_y[idxs]

        ax.scatter(gx, gy, s=dot_size, color=color, label=name,
                   edgecolors='white', linewidths=0.6, zorder=3)
        _pt_x.extend(gx.tolist())
        _pt_y.extend(gy.tolist())

        if show_labels and label_mode != 'none':
            for j, vi in enumerate(idxs):
                lbl = name if label_mode == 'custom' else (sample_names[vi] if vi < len(sample_names) else str(vi))
                _label_data.append((gx[j], gy[j], lbl, color))

        # 每组独立椭圆（虚线）
        if show_ellipse and len(idxs) >= 3:
            draw_hotelling_ellipse(np.column_stack([gx, gy]), color, lw=1.2, ls='--')

    # 原点十字轴
    ax.axhline(0, color='#555555', linewidth=0.8, zorder=1)
    ax.axvline(0, color='#555555', linewidth=0.8, zorder=1)

    if show_grid:
        ax.grid(True, linestyle='-', linewidth=0.5, color='white', zorder=0)

    # 轴范围
    if x_min is not None: ax.set_xlim(left=float(x_min))
    if x_max is not None: ax.set_xlim(right=float(x_max))
    if y_min is not None: ax.set_ylim(bottom=float(y_min))
    if y_max is not None: ax.set_ylim(top=float(y_max))

    xlabel = (xlabel_custom or f't[{pc_x}]')
    ylabel = (ylabel_custom or f't[{pc_y}]')
    ax.set_xlabel(xlabel, fontsize=axis_fontsize, color=axis_color, fontfamily=chinese_font or efont)
    ax.set_ylabel(ylabel, fontsize=axis_fontsize, color=axis_color, fontfamily=chinese_font or efont)
    ax.tick_params(labelsize=tick_fontsize, colors=tick_color)
    for tl in ax.get_xticklabels() + ax.get_yticklabels():
        tl.set_color(tick_color)

    if title:
        ax.set_title(title, fontsize=axis_fontsize + 1, fontfamily=chinese_font or efont)

    # 图例（右侧外部，始终显示）
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        leg = ax.legend(handles, labels, fontsize=legend_fontsize, frameon=True, framealpha=0.95,
                  edgecolor='#cccccc', loc='upper left',
                  bbox_to_anchor=(1.01, 1), borderaxespad=0,
                  prop={'family': chinese_font or efont, 'size': legend_fontsize})
        for text in leg.get_texts():
            text.set_fontfamily(chinese_font or efont)

    # 去掉上边框和右边框
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    # 标签放置：先渲染再用renderer计算bbox，迭代推开重叠
    if _label_data and show_labels and label_mode != 'none':
        fig.canvas.draw()
        renderer = fig.canvas.get_renderer()
        xlim = ax.get_xlim()
        ylim = ax.get_ylim()
        xspan = xlim[1] - xlim[0]
        yspan = ylim[1] - ylim[0]
        # 初始偏移：每个标签从自身圆点向外偏移一小段
        offset_x = xspan * 0.02
        offset_y = yspan * 0.02
        placed = []  # list of (txt_obj, px, py) where px,py = point coords
        for (px, py, lbl, color) in _label_data:
            # 初始方向：远离所有点的质心
            if len(_pt_x) > 1:
                cx = sum(_pt_x) / len(_pt_x)
                cy = sum(_pt_y) / len(_pt_y)
                dx = px - cx; dy = py - cy
                norm = (dx**2 + dy**2) ** 0.5 or 1
                tx = px + offset_x * dx / norm * 2
                ty = py + offset_y * dy / norm * 2
            else:
                tx, ty = px + offset_x, py + offset_y
            t = ax.text(tx, ty, lbl, fontsize=label_fontsize, color=color, zorder=4,
                        fontfamily=chinese_font or efont,
                        bbox=dict(boxstyle='round,pad=0.1', fc='white', ec='none', alpha=0.7))
            placed.append((t, px, py))

        # 迭代推开：标签bbox只排斥自身对应圆点 + 其他标签bbox
        for _ in range(200):
            moved = False
            fig.canvas.draw()
            bboxes = [t.get_window_extent(renderer) for t, _, _ in placed]
            for i, (t, px, py) in enumerate(placed):
                bb = bboxes[i]
                fx, fy = 0.0, 0.0
                # 只排斥自身对应圆点
                qd = ax.transData.transform((px, py))
                dx = bb.x0 + bb.width/2 - qd[0]
                dy = bb.y0 + bb.height/2 - qd[1]
                dist = (dx**2 + dy**2) ** 0.5 or 1
                overlap = (bb.width/2 + 14) - dist
                if overlap > 0:
                    fx += dx / dist * overlap * 0.5
                    fy += dy / dist * overlap * 0.5
                # 排斥：与其他文字bbox（基于实际重叠量，力度更强）
                for j, (t2, _, _) in enumerate(placed):
                    if i == j: continue
                    bb2 = bboxes[j]
                    ox = min(bb.x1, bb2.x1) - max(bb.x0, bb2.x0)
                    oy = min(bb.y1, bb2.y1) - max(bb.y0, bb2.y0)
                    if ox > 0 and oy > 0:
                        dx = (bb.x0+bb.width/2) - (bb2.x0+bb2.width/2)
                        dy = (bb.y0+bb.height/2) - (bb2.y0+bb2.height/2)
                        dist = (dx**2+dy**2)**0.5 or 1
                        # 沿最短轴推开
                        if ox < oy:
                            fx += (1 if dx >= 0 else -1) * ox * 0.6
                        else:
                            fy += (1 if dy >= 0 else -1) * oy * 0.6
                if abs(fx) > 0.1 or abs(fy) > 0.1:
                    cur = ax.transData.transform(t.get_position())
                    new_disp = (cur[0]+fx, cur[1]+fy)
                    new_data = ax.transData.inverted().transform(new_disp)
                    new_data[0] = max(xlim[0], min(xlim[1], new_data[0]))
                    new_data[1] = max(ylim[0], min(ylim[1], new_data[1]))
                    t.set_position(new_data)
                    moved = True
            if not moved:
                break

    # Footer
    footer = f"R2X[{pc_x}] = {r2x[pc_x-1]:.4f}    R2X[{pc_y}] = {r2x[pc_y-1]:.4f}"
    if show_ellipse:
        footer += "    Ellipse: Hotelling's T2 (95%)"
    ax.text(0.5, -0.10, footer, transform=ax.transAxes, ha='center', va='top',
            fontsize=tick_fontsize - 1, fontfamily=efont, color='#444444')

    fig.subplots_adjust(bottom=0.12, right=0.78)

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    img_b64 = base64.b64encode(buf.read()).decode()

    score_df = pd.DataFrame(scores, columns=[f't[{i+1}]' for i in range(n_components)])
    excel_buf = io.BytesIO()
    score_df.to_excel(excel_buf, index=False)
    excel_buf.seek(0)
    excel_b64 = base64.b64encode(excel_buf.read()).decode()

    return img_b64, excel_b64


if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, use_reloader=False, host='0.0.0.0', port=port)
